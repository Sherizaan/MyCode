from openpyxl import load_workbook
class start:
	def __init__(self):
		self.movie_selected=""
		self.movie_list={}
		self.open_cinema()

	def book_seat(self):
		for seats in self.sheet.iter_rows(min_row=1,min_col=1,max_col=5,values_only=True):
		 	print(seats)
		self.interface()

	def select_movie(self):
		print("movie options")
		for key,value in self.movie_list.items():
			print("\t{}. {}".format(key,value))
		while(True):
			choice=int(input("option: "))
			if choice <1 or choice > len(self.movie_list):
				print("invalid selection")
			else:
				self.movie_selected = self.movie_list["{}".format(choice)]
				self.sheet = self.workbook["{}".format(self.movie_selected)]
				break
		self.interface()
		
	def interface(self):
		print("movie_selected: {}".format(self.movie_selected))
		print("Menue options")
		if self.movie_selected!="":
			print("\t1. select a different movie")
		else:
			print("\t1. select a movie")
		print("\t2. book a seat")
		print("\t3. exit")
		choice=int(input("Option: "))
		if choice < 1 or choice > 3:
			print("invalid selection\n")
			self.interface()
		if choice == 1:
			self.select_movie()
		elif choice == 2:
			self.book_seat()
		elif choice == 3:
			exit()

	def open_cinema(self):
		self.workbook=load_workbook(filename="Cinemas.xlsx")
		count=1
		for x in self.workbook.sheetnames:
			self.movie_list["{}".format(count)]=x
			count+=1
		
		

run=start()
run.interface()
